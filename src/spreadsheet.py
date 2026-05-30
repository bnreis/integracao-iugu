"""
Gerenciamento da planilha Excel de empresas cadastradas.

A planilha contém as empresas cadastradas para:
1. Receberem NFS-e automaticamente quando pagam uma fatura na Iugu
2. Terem boletos gerados automaticamente no dia do mês configurado

Endereço do tomador: NÃO fica na planilha. Para NFS-e, é obtido da própria
fatura da Iugu (payload do webhook). Para boleto recorrente, a Iugu usa o
cadastro do customer já existente na conta.

Colunas:
- cnpj (chave de busca — só números, 14 dígitos)
- razao_social, email
- NFS-e:
    - codigo_servico, descricao_servico, aliquota_iss
    - emitir_nf: True = emite NFS-e automaticamente quando fatura é paga
- BOLETO RECORRENTE:
    - descricao_boleto: texto que aparece na cobrança do boleto
    - valor_fatura: valor em reais (ex: "1850,00"); vazio/0 = não gera boleto
    - dia_criacao_fatura: dia do mês (1-31) em que o boleto é criado
- observacoes, ativo (True/False — desabilita tudo sem remover a linha)
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import settings

# Colunas da planilha (ordem e nomes canônicos)
COLUNAS = [
    "cnpj",
    "razao_social",
    "email",
    "codigo_servico",
    "descricao_servico",
    "aliquota_iss",
    "emitir_nf",
    "nf_na_criacao",
    "descricao_boleto",
    "valor_fatura",
    "dia_criacao_fatura",
    "observacoes",
    "ativo",
]


@dataclass
class Empresa:
    """Representa uma empresa cadastrada no sistema (NFS-e + boleto recorrente)."""

    cnpj: str
    razao_social: str = ""
    email: str = ""
    codigo_servico: str = ""
    descricao_servico: str = ""
    aliquota_iss: float = 0.0
    emitir_nf: bool = True
    nf_na_criacao: bool = False     # True = emite NFS-e junto com a fatura (não espera pgto)
    descricao_boleto: str = ""
    valor_fatura: str = ""          # formato BR: "1850,00" (vazio = sem cobrança auto)
    dia_criacao_fatura: int = 0     # 1-31 (0 = desabilitado)
    observacoes: str = ""
    ativo: bool = True

    def to_dict(self) -> dict:
        return {c: getattr(self, c) for c in COLUNAS}

    @property
    def valor_fatura_cents(self) -> int:
        """Converte 'valor_fatura' (BR: '1850,00' ou '1.850,00') para centavos (int)."""
        return parse_valor_br_to_cents(self.valor_fatura)

    def tem_boleto_recorrente(self) -> bool:
        """True se a empresa tem valor e dia configurados para cobrança automática."""
        return self.valor_fatura_cents > 0 and 1 <= self.dia_criacao_fatura <= 31


def parse_valor_br_to_cents(valor) -> int:
    """
    Converte valor no formato BR para centavos.

    Aceita:
        "1850,00"       → 185000
        "1.850,00"      → 185000
        "R$ 1.850,00"   → 185000
        1850.00 (float) → 185000
        0, "", None     → 0
    """
    if valor is None or valor == "":
        return 0
    if isinstance(valor, (int, float)):
        return int(round(float(valor) * 100))
    s = str(valor).strip()
    if not s:
        return 0
    # Remove símbolos de moeda e espaços
    s = s.replace("R$", "").replace("r$", "").replace(" ", "").strip()
    # Remove separador de milhar e normaliza vírgula decimal
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return int(round(float(s) * 100))
    except ValueError:
        logger.warning(f"Valor inválido ao converter: {valor!r}")
        return 0


def format_cents_to_br(cents: int) -> str:
    """Formato inverso — útil para logs/relatórios. 185000 → '1.850,00'."""
    if cents <= 0:
        return "0,00"
    reais, centavos = divmod(cents, 100)
    # Insere separador de milhar nos reais
    reais_str = f"{reais:,}".replace(",", ".")
    return f"{reais_str},{centavos:02d}"


def normalizar_cnpj(cnpj: str) -> str:
    """Remove tudo que não é dígito e retorna apenas números."""
    if not cnpj:
        return ""
    return "".join(filter(str.isdigit, str(cnpj)))


def criar_planilha_modelo(caminho: Path | None = None, com_exemplos: bool = True) -> Path:
    """
    Cria uma planilha modelo em branco (com ou sem linhas de exemplo).

    Args:
        caminho: onde salvar (default: settings.planilha_empresas)
        com_exemplos: se True, inclui duas linhas fictícias de exemplo

    Returns:
        Caminho absoluto do arquivo criado.
    """
    caminho = Path(caminho) if caminho else settings.planilha_empresas

    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas Autorizadas"

    # Cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for idx, coluna in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=idx, value=coluna)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Largura das colunas (aproximação razoável)
    larguras = {
        "cnpj": 20, "razao_social": 35, "email": 30,
        "codigo_servico": 15, "descricao_servico": 35, "aliquota_iss": 15,
        "emitir_nf": 12, "nf_na_criacao": 15,
        "descricao_boleto": 35, "valor_fatura": 15, "dia_criacao_fatura": 18,
        "observacoes": 30, "ativo": 10,
    }
    for idx, coluna in enumerate(COLUNAS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = larguras.get(coluna, 15)

    # Congelar cabeçalho
    ws.freeze_panes = "A2"

    # Exemplos
    if com_exemplos:
        exemplos = [
            Empresa(
                cnpj="12345678000190",
                razao_social="Empresa Exemplo LTDA",
                email="financeiro@exemplo.com.br",
                codigo_servico="01.07",
                descricao_servico="Suporte técnico em informática",
                aliquota_iss=2.0,
                emitir_nf=True,
                descricao_boleto="Mensalidade suporte técnico",
                valor_fatura="1850,00",
                dia_criacao_fatura=10,
                observacoes="Cliente mensal",
                ativo=True,
            ),
            Empresa(
                cnpj="98765432000155",
                razao_social="Outra Empresa SA",
                email="notas@outraempresa.com",
                codigo_servico="17.11",
                descricao_servico="Administração em geral",
                aliquota_iss=5.0,
                emitir_nf=True,
                descricao_boleto="Assessoria administrativa mensal",
                valor_fatura="3500,00",
                dia_criacao_fatura=25,
                observacoes="",
                ativo=True,
            ),
        ]
        for row_idx, emp in enumerate(exemplos, start=2):
            for col_idx, coluna in enumerate(COLUNAS, start=1):
                valor = getattr(emp, coluna)
                # Padroniza: booleanos → "Sim"/"Não", valor_fatura → "R$ X.XXX,XX"
                if isinstance(valor, bool):
                    valor = "Sim" if valor else "Não"
                if coluna == "valor_fatura" and valor:
                    cents = parse_valor_br_to_cents(valor)
                    if cents > 0:
                        valor = f"R$ {format_cents_to_br(cents)}"
                ws.cell(row=row_idx, column=col_idx, value=valor)

    # Aba de instruções
    ws_info = wb.create_sheet("Instruções")
    instrucoes = [
        ("COMO USAR ESTA PLANILHA", True),
        ("", False),
        ("1. Preencha uma linha por empresa cadastrada.", False),
        ("2. O CNPJ é a chave de busca — pode ser digitado com ou sem formatação.", False),
        ("3. Endereço NÃO é mais necessário aqui — é obtido da Iugu automaticamente.", False),
        ("", False),
        ("CAMPOS DE NFS-e:", True),
        ("4. codigo_servico: consulte a lista do DF em https://iss.fazenda.df.gov.br/online", False),
        ("5. aliquota_iss: percentual (ex: 2.0 para 2%)", False),
        ("6. emitir_nf: True = emite NFS-e automaticamente ao receber pagamento.", False),
        ("   False = ignora emissão automática (cliente paga mas NFS-e não é gerada).", False),
        ("6b. nf_na_criacao: True = emite NFS-e junto com a fatura (na criação do boleto),", False),
        ("   e NÃO emite novamente quando o pagamento é recebido.", False),
        ("   False (padrão) = NFS-e só é emitida após o pagamento.", False),
        ("", False),
        ("CAMPOS DE BOLETO RECORRENTE:", True),
        ("7. descricao_boleto: texto que aparece na cobrança da Iugu.", False),
        ("   (ex: 'Mensalidade de suporte técnico — Abril/2026')", False),
        ("8. valor_fatura: valor em reais no formato brasileiro.", False),
        ("   Exemplos válidos: 1850,00 | 1.850,00 | R$ 1.850,00", False),
        ("9. dia_criacao_fatura: dia do mês (1-31) em que o boleto será criado.", False),
        ("   - Se o mês não tiver o dia (ex: 31 em fevereiro), usa o último dia disponível.", False),
        ("   - Vazio ou 0 = NÃO gera cobrança automática (só NFS-e se emitir_nf=True).", False),
        ("   - Vencimento automático: 10 dias após a criação.", False),
        ("", False),
        ("CONTROLE GERAL:", True),
        ("10. ativo: 'False' desabilita TUDO (NFS-e e boleto recorrente).", False),
        ("", False),
        ("ATENÇÃO:", True),
        ("- Não altere os nomes das colunas (linha 1).", False),
        ("- Não altere a ordem das colunas.", False),
        ("- CNPJs duplicados: será usada a primeira ocorrência ativa.", False),
        ("- Para desabilitar APENAS o boleto recorrente (mantendo NFS-e),", False),
        ("  deixe valor_fatura e dia_criacao_fatura em branco.", False),
        ("- Para desabilitar APENAS a NFS-e (mantendo cobrança), coloque emitir_nf=False.", False),
    ]
    for idx, (texto, negrito) in enumerate(instrucoes, start=1):
        cell = ws_info.cell(row=idx, column=1, value=texto)
        if negrito:
            cell.font = Font(bold=True, size=12)
    ws_info.column_dimensions["A"].width = 90

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    logger.info(f"Planilha modelo criada em {caminho}")
    return caminho


class EmpresasRepository:
    """Lê e consulta a planilha de empresas autorizadas."""

    def __init__(self, caminho: Path | None = None):
        self.caminho = Path(caminho) if caminho else settings.planilha_empresas
        self._empresas: dict[str, Empresa] = {}
        self._carregada = False

    def carregar(self, forcar: bool = False) -> None:
        """Carrega a planilha em memória, indexada por CNPJ normalizado."""
        if self._carregada and not forcar:
            return

        if not self.caminho.exists():
            raise FileNotFoundError(
                f"Planilha não encontrada em {self.caminho}. "
                f"Execute scripts/create_spreadsheet.py para criar uma."
            )

        wb = load_workbook(self.caminho, data_only=True)
        ws = wb["Empresas Autorizadas"] if "Empresas Autorizadas" in wb.sheetnames else wb.active

        # Lê o cabeçalho para mapear colunas
        header = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                header[str(cell.value).strip().lower()] = idx

        # Valida colunas MÍNIMAS obrigatórias.
        # Colunas de endereço foram REMOVIDAS (endereço agora vem da Iugu).
        # Retrocompatibilidade: colunas novas (emitir_nf, descricao_boleto,
        # valor_fatura, dia_criacao_fatura) podem estar ausentes em planilhas antigas.
        COLUNAS_OBRIGATORIAS = {
            "cnpj", "razao_social", "email",
            "codigo_servico", "descricao_servico", "aliquota_iss",
            "observacoes", "ativo",
        }
        faltantes = [c for c in COLUNAS_OBRIGATORIAS if c not in header]
        if faltantes:
            raise ValueError(
                f"Planilha inválida — colunas faltantes: {faltantes}. "
                f"Esperado no mínimo: {sorted(COLUNAS_OBRIGATORIAS)}"
            )

        # Avisa se colunas novas faltarem (só em warning, não quebra)
        novas = {"emitir_nf", "nf_na_criacao", "descricao_boleto", "valor_fatura", "dia_criacao_fatura"}
        faltam_novas = [c for c in novas if c not in header]
        if faltam_novas:
            logger.warning(
                f"Planilha sem colunas novas: {faltam_novas}. "
                f"Execute scripts/migrate_simplify_schema.py para atualizá-las."
            )

        def _valor_coluna(row_, nome: str, default=None):
            """Retorna o valor da coluna se existir no header, senão o default."""
            idx = header.get(nome)
            if idx is None:
                return default
            return row_[idx - 1]

        self._empresas.clear()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not _valor_coluna(row, "cnpj"):
                continue

            cnpj = normalizar_cnpj(_valor_coluna(row, "cnpj"))
            if not cnpj:
                continue

            # Helper local para converter valores boolean-like
            # Aceita: True/False (bool), "True"/"False" (str), "Sim"/"Não" (str),
            # "VERDADEIRO"/"FALSO" (Excel pt-BR), "1"/"0", "yes"/"no"
            def _to_bool(valor, padrao: bool = False) -> bool:
                if valor is None or valor == "":
                    return padrao
                if isinstance(valor, str):
                    return valor.strip().lower() in (
                        "true", "sim", "s", "1", "yes", "y", "verdadeiro",
                    )
                return bool(valor)

            # Converte ativo: aceita True/False/"sim"/"não"/1/0
            ativo = _to_bool(_valor_coluna(row, "ativo", True), padrao=True)

            # emitir_nf: default True (retrocompatível — se coluna não existe, emite)
            emitir_nf = _to_bool(_valor_coluna(row, "emitir_nf", True), padrao=True)

            # nf_na_criacao: default False (retrocompatível — se coluna não existe, não emite na criação)
            nf_na_criacao = _to_bool(_valor_coluna(row, "nf_na_criacao", False), padrao=False)

            # Dia de criação da fatura: inteiro 1-31 ou 0
            dia_raw = _valor_coluna(row, "dia_criacao_fatura", 0)
            try:
                dia_criacao = int(float(dia_raw)) if dia_raw not in (None, "") else 0
            except (ValueError, TypeError):
                dia_criacao = 0
            if not (0 <= dia_criacao <= 31):
                logger.warning(
                    f"CNPJ {cnpj}: dia_criacao_fatura fora do intervalo (1-31): {dia_raw!r}"
                )
                dia_criacao = 0

            # Aliquota ISS
            try:
                aliq = float(_valor_coluna(row, "aliquota_iss", 0) or 0)
            except (ValueError, TypeError):
                aliq = 0.0

            def _str(campo: str) -> str:
                v = _valor_coluna(row, campo, "")
                return str(v).strip() if v is not None else ""

            empresa = Empresa(
                cnpj=cnpj,
                razao_social=_str("razao_social"),
                email=_str("email"),
                codigo_servico=_str("codigo_servico"),
                descricao_servico=_str("descricao_servico"),
                aliquota_iss=aliq,
                emitir_nf=emitir_nf,
                nf_na_criacao=nf_na_criacao,
                descricao_boleto=_str("descricao_boleto"),
                valor_fatura=_str("valor_fatura"),
                dia_criacao_fatura=dia_criacao,
                observacoes=_str("observacoes"),
                ativo=ativo,
            )
            self._empresas[cnpj] = empresa

        self._carregada = True
        logger.info(f"Planilha carregada: {len(self._empresas)} empresas")

    def empresas_com_boleto_recorrente(self) -> list[Empresa]:
        """Retorna empresas ativas que têm valor e dia configurados para cobrança automática."""
        self.carregar()
        return [e for e in self._empresas.values() if e.ativo and e.tem_boleto_recorrente()]

    def buscar_por_cnpj(self, cnpj: str) -> Optional[Empresa]:
        """Retorna a empresa pelo CNPJ (só empresas ativas). None se não encontrada."""
        self.carregar()
        empresa = self._empresas.get(normalizar_cnpj(cnpj))
        if empresa and empresa.ativo:
            return empresa
        return None

    def listar_ativas(self) -> list[Empresa]:
        """Retorna todas as empresas ativas."""
        self.carregar()
        return [e for e in self._empresas.values() if e.ativo]

    def existe(self, cnpj: str) -> bool:
        """Verifica se um CNPJ está autorizado (e ativo) para emissão automática."""
        return self.buscar_por_cnpj(cnpj) is not None
