"""
Migra a planilha existente adicionando as 3 colunas novas de boleto recorrente:
- descricao_boleto
- valor_fatura
- dia_criacao_fatura

É IDEMPOTENTE: se as colunas já existem, não faz nada.
PRESERVA todos os dados atuais (CNPJs, endereços, NFS-e) da planilha.

Uso:
    # Backup automático, migração + atualização da aba Instruções
    python scripts/migrate_add_billing_columns.py

    # Dry-run (só mostra o que faria, sem modificar)
    python scripts/migrate_add_billing_columns.py --dry-run

    # Migrar um arquivo diferente do padrão
    python scripts/migrate_add_billing_columns.py --arquivo outra_planilha.xlsx

⚠️ IMPORTANTE: feche o Excel antes de rodar (se o arquivo estiver aberto).
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import settings

# Colunas NOVAS que serão adicionadas (em ordem, antes de "observacoes")
NOVAS_COLUNAS = ["descricao_boleto", "valor_fatura", "dia_criacao_fatura"]

# Posição alvo: depois de "aliquota_iss" e antes de "observacoes"
COLUNA_ANCORA_ANTES = "aliquota_iss"

# Larguras padrão para as colunas novas
LARGURAS = {
    "descricao_boleto": 35,
    "valor_fatura": 15,
    "dia_criacao_fatura": 18,
}


def fazer_backup(caminho: Path) -> Path:
    """Cria uma cópia de backup com timestamp antes de mexer."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = caminho.with_name(f"{caminho.stem}.bak_{ts}{caminho.suffix}")
    shutil.copy2(caminho, backup)
    logger.info(f"Backup criado: {backup.name}")
    return backup


def migrar(caminho: Path, dry_run: bool = False) -> dict:
    """Aplica a migração. Retorna um dict com o que foi feito."""
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    wb = load_workbook(caminho)
    if "Empresas Autorizadas" not in wb.sheetnames:
        raise ValueError(
            f'Aba "Empresas Autorizadas" não encontrada. '
            f'Abas disponíveis: {wb.sheetnames}'
        )

    ws = wb["Empresas Autorizadas"]

    # Mapeia colunas existentes pelo cabeçalho
    header_atual = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            header_atual[str(cell.value).strip().lower()] = col_idx

    ja_existentes = [c for c in NOVAS_COLUNAS if c in header_atual]
    faltantes = [c for c in NOVAS_COLUNAS if c not in header_atual]

    if not faltantes:
        logger.info("✅ Todas as colunas novas já existem. Nada a fazer.")
        return {"status": "ja_migrada", "colunas_existentes": ja_existentes}

    # Posição onde inserir — logo depois da âncora
    ancora_idx = header_atual.get(COLUNA_ANCORA_ANTES)
    if ancora_idx is None:
        raise ValueError(
            f'Coluna de ancoragem "{COLUNA_ANCORA_ANTES}" não encontrada. '
            f'Esta planilha não parece ser do projeto.'
        )
    pos_insercao = ancora_idx + 1
    logger.info(
        f"Inserindo {len(faltantes)} coluna(s) a partir da posição "
        f"{get_column_letter(pos_insercao)} (coluna {pos_insercao})"
    )

    if dry_run:
        logger.info("DRY RUN — nenhuma alteração será feita.")
        print(f"\nColunas que seriam adicionadas: {faltantes}")
        print(f"Posição de inserção: coluna {get_column_letter(pos_insercao)}")
        print(f"Total de linhas de dados: {ws.max_row - 1}")
        return {"status": "dry_run", "colunas_adicionadas": faltantes}

    # Backup antes de mexer
    fazer_backup(caminho)

    # Insere as colunas novas (openpyxl insere da esquerda para direita)
    ws.insert_cols(pos_insercao, amount=len(faltantes))

    # Estilo do cabeçalho (mesmo padrão do modelo original)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    for offset, nome_coluna in enumerate(faltantes):
        col = pos_insercao + offset
        cell = ws.cell(row=1, column=col, value=nome_coluna)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = LARGURAS.get(nome_coluna, 18)

    # Atualiza/recria aba de instruções para refletir as colunas novas
    _atualizar_aba_instrucoes(wb)

    wb.save(caminho)
    logger.info(f"✅ Planilha migrada: {caminho}")
    return {
        "status": "migrada",
        "colunas_adicionadas": faltantes,
        "colunas_ja_existentes": ja_existentes,
        "linhas_dados": ws.max_row - 1,
    }


def _atualizar_aba_instrucoes(wb) -> None:
    """Recria a aba 'Instruções' com o texto atualizado."""
    if "Instruções" in wb.sheetnames:
        del wb["Instruções"]
    ws_info = wb.create_sheet("Instruções")
    instrucoes = [
        ("COMO USAR ESTA PLANILHA", True),
        ("", False),
        ("1. Preencha uma linha por empresa cadastrada.", False),
        ("2. O CNPJ é a chave de busca — com ou sem formatação.", False),
        ("", False),
        ("CAMPOS DE NFS-e:", True),
        ("3. Endereço completo é obrigatório para emissão da NFS-e.", False),
        ("4. codigo_servico: https://iss.fazenda.df.gov.br/online", False),
        ("5. aliquota_iss: percentual (ex: 2.0 para 2%)", False),
        ("", False),
        ("CAMPOS DE BOLETO RECORRENTE (NOVOS):", True),
        ("6. descricao_boleto: texto que aparece no boleto da Iugu.", False),
        ("7. valor_fatura: formato BR (1850,00). Vazio = sem cobrança automática.", False),
        ("8. dia_criacao_fatura: dia do mês (1-31).", False),
        ("   - Se o mês não tiver o dia, usa o último dia disponível.", False),
        ("   - Vencimento: 10 dias após criação.", False),
        ("", False),
        ("CONTROLE:", True),
        ("9. ativo: 'False' desabilita TUDO.", False),
        ("   Para desligar APENAS o boleto recorrente mantendo NFS-e,", False),
        ("   deixe valor_fatura e dia_criacao_fatura em branco.", False),
    ]
    for idx, (texto, negrito) in enumerate(instrucoes, start=1):
        cell = ws_info.cell(row=idx, column=1, value=texto)
        if negrito:
            cell.font = Font(bold=True, size=12)
    ws_info.column_dimensions["A"].width = 90


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument(
        "--arquivo",
        type=Path,
        default=None,
        help=f"Caminho da planilha (default: {settings.planilha_empresas})",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    caminho = args.arquivo or settings.planilha_empresas
    logger.info(f"Planilha alvo: {caminho}")

    try:
        resultado = migrar(caminho, dry_run=args.dry_run)
    except Exception as exc:
        logger.exception("Falha na migração")
        print(f"\n❌ ERRO: {exc}")
        sys.exit(1)

    print(f"\n{'=' * 60}")
    print(f"📊 MIGRAÇÃO — status: {resultado['status'].upper()}")
    print(f"{'=' * 60}")
    if resultado["status"] == "migrada":
        print(f"Colunas adicionadas: {resultado['colunas_adicionadas']}")
        print(f"Linhas de dados preservadas: {resultado['linhas_dados']}")
        print(f"\n✅ Abra a planilha no Excel e preencha as novas colunas.")
    elif resultado["status"] == "ja_migrada":
        print(f"Nada a fazer. Colunas já existentes: {resultado['colunas_existentes']}")
    elif resultado["status"] == "dry_run":
        print(f"Seriam adicionadas: {resultado['colunas_adicionadas']}")


if __name__ == "__main__":
    main()
