"""
Adiciona a coluna `nf_na_criacao` à planilha de empresas.

Essa coluna controla se a NFS-e deve ser emitida no momento da criação
da fatura (junto com o boleto), em vez de esperar o pagamento.

IDEMPOTENTE: pode ser rodado várias vezes sem problema.

Uso:
    python scripts/migrate_add_nf_na_criacao.py
    python scripts/migrate_add_nf_na_criacao.py --dry-run
    python scripts/migrate_add_nf_na_criacao.py --empresas-true "10519719000197,33498643000166,04955204000137"

⚠️ Feche o Excel antes de rodar.
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import settings

COLUNA_NOVA = "nf_na_criacao"
# Posição alvo: logo depois de "emitir_nf"
COLUNA_ANCORA = "emitir_nf"


def fazer_backup(caminho: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = caminho.with_name(f"{caminho.stem}.bak_{ts}{caminho.suffix}")
    shutil.copy2(caminho, backup)
    logger.info(f"Backup criado: {backup.name}")
    return backup


def _mapear_header(ws):
    m = {}
    for ci, c in enumerate(ws[1], start=1):
        if c.value:
            m[str(c.value).strip().lower()] = ci
    return m


def normalizar_cnpj(cnpj: str) -> str:
    return "".join(c for c in str(cnpj) if c.isdigit())


def migrar(
    caminho: Path,
    dry_run: bool = False,
    empresas_true: list[str] | None = None,
) -> dict:
    """Aplica a migração. Retorna dict com o que foi feito."""
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    wb = load_workbook(caminho)

    # Procura a aba correta
    aba = None
    for nome_aba in wb.sheetnames:
        if "empresa" in nome_aba.lower():
            aba = nome_aba
            break
    if aba is None:
        aba = wb.sheetnames[0]
    ws = wb[aba]

    header = _mapear_header(ws)
    total_dados = ws.max_row - 1

    if COLUNA_NOVA in header:
        logger.info(f"✅ Coluna '{COLUNA_NOVA}' já existe na posição {header[COLUNA_NOVA]}. Nada a fazer.")

        # Se --empresas-true foi passado, atualiza os valores mesmo que a coluna exista
        if empresas_true:
            cnpjs_true = set(normalizar_cnpj(c) for c in empresas_true)
            col_nf = header[COLUNA_NOVA]
            col_cnpj = header.get("cnpj")
            atualizados = []
            for r in range(2, ws.max_row + 1):
                cnpj_val = ws.cell(row=r, column=col_cnpj).value
                if cnpj_val:
                    cnpj_norm = normalizar_cnpj(cnpj_val)
                    if cnpj_norm in cnpjs_true:
                        ws.cell(row=r, column=col_nf, value="Sim")
                        atualizados.append(cnpj_norm)
                        logger.info(f"  ✔ CNPJ {cnpj_norm} → nf_na_criacao=Sim")
            if atualizados and not dry_run:
                wb.save(caminho)
                logger.info(f"Planilha salva com {len(atualizados)} empresas atualizadas")
            return {"status": "atualizada", "atualizados": atualizados}

        return {"status": "ja_migrada", "linhas_dados": total_dados}

    if dry_run:
        logger.info("DRY RUN — nenhuma alteração será feita.")
        return {"status": "dry_run", "linhas_dados": total_dados}

    # Backup
    fazer_backup(caminho)

    # Encontra posição para inserir (depois de emitir_nf)
    ancora = header.get(COLUNA_ANCORA)
    if ancora is None:
        # Se emitir_nf não existe, coloca depois de aliquota_iss
        ancora = header.get("aliquota_iss")
    if ancora is None:
        # Fallback: coloca no final
        ancora = ws.max_column

    pos = ancora + 1
    ws.insert_cols(pos, amount=1)

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell = ws.cell(row=1, column=pos, value=COLUNA_NOVA)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(pos)].width = 15

    # CNPJs que devem ter nf_na_criacao=True
    cnpjs_true = set()
    if empresas_true:
        cnpjs_true = set(normalizar_cnpj(c) for c in empresas_true)

    # Preenche default para todas as linhas de dados
    header_atualizado = _mapear_header(ws)
    col_cnpj = header_atualizado.get("cnpj")
    count_true = 0
    for r in range(2, ws.max_row + 1):
        cnpj_val = ws.cell(row=r, column=col_cnpj).value if col_cnpj else None
        if cnpj_val:
            cnpj_norm = normalizar_cnpj(cnpj_val)
            if cnpj_norm in cnpjs_true:
                ws.cell(row=r, column=pos, value="Sim")
                count_true += 1
                logger.info(f"  ✔ CNPJ {cnpj_norm} → nf_na_criacao=Sim")
            else:
                ws.cell(row=r, column=pos, value="Não")

    wb.save(caminho)
    logger.info(f"✅ Planilha migrada: coluna '{COLUNA_NOVA}' adicionada na posição {get_column_letter(pos)}")

    return {
        "status": "migrada",
        "posicao": get_column_letter(pos),
        "linhas_dados": total_dados,
        "empresas_true": count_true,
    }


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument(
        "--arquivo", type=Path, default=None,
        help=f"Caminho da planilha (default: {settings.planilha_empresas})",
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--empresas-true", type=str, default=None,
        help="CNPJs separados por vírgula que devem ter nf_na_criacao=True",
    )
    args = parser.parse_args()

    caminho = args.arquivo or settings.planilha_empresas
    logger.info(f"Planilha alvo: {caminho}")

    empresas_true = None
    if args.empresas_true:
        empresas_true = [c.strip() for c in args.empresas_true.split(",") if c.strip()]

    try:
        resultado = migrar(caminho, dry_run=args.dry_run, empresas_true=empresas_true)
    except Exception as exc:
        logger.exception("Falha na migração")
        print(f"\n❌ ERRO: {exc}")
        sys.exit(1)

    print(f"\n{'=' * 60}")
    print(f"📊 MIGRAÇÃO nf_na_criacao — status: {resultado['status'].upper()}")
    print(f"{'=' * 60}")
    for k, v in resultado.items():
        if k != "status":
            print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
