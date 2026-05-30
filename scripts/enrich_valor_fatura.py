"""
Enriquece a planilha preenchendo `valor_fatura` (e opcionalmente `descricao_boleto`)
com base na ÚLTIMA fatura de cada CNPJ na Iugu.

Estratégia: varre as faturas da Iugu dos últimos N dias, em ordem decrescente,
e para cada CNPJ listado na planilha captura o valor da fatura mais recente.

Uso:
    # Busca faturas dos últimos 120 dias e preenche valor_fatura
    python scripts/enrich_valor_fatura.py

    # Janela maior (para clientes com faturas mais antigas)
    python scripts/enrich_valor_fatura.py --days 180

    # Preenche também a descrição do boleto com a descrição do item da última fatura
    python scripts/enrich_valor_fatura.py --preencher-descricao

    # Dry-run (mostra o que faria, não salva)
    python scripts/enrich_valor_fatura.py --dry-run

    # Preserva valores já preenchidos na planilha (default: sobrescreve se vier valor da Iugu)
    python scripts/enrich_valor_fatura.py --preservar

Atenção:
    - Feche o Excel antes de rodar (senão openpyxl não consegue escrever)
    - Cria backup automático antes de sobrescrever
"""
from __future__ import annotations

import argparse
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from loguru import logger
from openpyxl import load_workbook

from src.config import settings
from src.iugu_client import IuguClient
from src.spreadsheet import (
    COLUNAS,
    EmpresasRepository,
    format_cents_to_br,
    normalizar_cnpj,
)

CNPJ_LENGTH = 14


def _backup(caminho: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = caminho.with_name(f"{caminho.stem}.bak_{ts}{caminho.suffix}")
    shutil.copy2(caminho, destino)
    logger.info(f"Backup salvo em {destino.name}")
    return destino


def _parse_iso_dt(s: str | None) -> datetime:
    """Parseia datetime ISO com fallback. Usado para ordenação."""
    if not s:
        return datetime.min
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return datetime.min


def buscar_ultima_fatura_por_cnpj(
    cnpjs_alvo: set[str],
    dias: int = 120,
) -> dict[str, dict]:
    """
    Retorna um dict {cnpj: fatura_mais_recente} percorrendo TODAS as faturas
    dos últimos `dias` dias (paginado), em qualquer status.
    """
    data_fim = date.today()
    data_inicio = data_fim - timedelta(days=dias)
    logger.info(f"Buscando faturas de {data_inicio} a {data_fim}")

    melhor_por_cnpj: dict[str, dict] = {}
    melhor_dt: dict[str, datetime] = {}

    with IuguClient() as client:
        start = 0
        limit = 100
        pagina = 0
        total_visto = 0
        while True:
            pagina += 1
            result = client.list_invoices(
                limit=limit,
                start=start,
                created_at_from=data_inicio.isoformat(),
                created_at_to=data_fim.isoformat(),
                sortBy="created_at",
            )
            items = result.get("items", []) or []
            total = result.get("totalItems", 0)
            logger.info(f"  Página {pagina}: {len(items)} faturas (de {total} total)")

            for inv in items:
                total_visto += 1
                doc_raw = inv.get("payer_cpf_cnpj") or ""
                cnpj = normalizar_cnpj(doc_raw)
                if len(cnpj) != CNPJ_LENGTH:
                    continue
                if cnpj not in cnpjs_alvo:
                    continue
                dt_fatura = _parse_iso_dt(inv.get("created_at_iso"))
                if cnpj not in melhor_dt or dt_fatura > melhor_dt[cnpj]:
                    melhor_por_cnpj[cnpj] = inv
                    melhor_dt[cnpj] = dt_fatura

            if len(items) < limit:
                break
            start += limit

    logger.info(
        f"✔ {total_visto} faturas varridas, "
        f"{len(melhor_por_cnpj)}/{len(cnpjs_alvo)} CNPJs encontrados na Iugu"
    )
    return melhor_por_cnpj


def _pegar_descricao_item(inv: dict) -> str:
    items = inv.get("items") or []
    if items:
        desc = items[0].get("description") or ""
        return str(desc).strip()
    return ""


def atualizar_planilha(
    caminho: Path,
    valores_por_cnpj: dict[str, dict],
    preservar_existentes: bool,
    preencher_descricao: bool,
    dry_run: bool,
) -> dict:
    """Grava os valores novos na planilha. Retorna estatísticas do que foi feito."""
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    if not dry_run:
        _backup(caminho)

    wb = load_workbook(caminho)
    ws = wb["Empresas Autorizadas"]

    # Mapeia colunas pelo header
    header = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            header[str(cell.value).strip().lower()] = idx

    col_cnpj = header.get("cnpj")
    col_valor = header.get("valor_fatura")
    col_desc_boleto = header.get("descricao_boleto")

    if not col_cnpj or not col_valor:
        raise ValueError(
            "Planilha sem colunas 'cnpj' e/ou 'valor_fatura'. "
            "Execute scripts/migrate_add_billing_columns.py primeiro."
        )

    stats = defaultdict(int)
    alteracoes_log = []

    for row_idx in range(2, ws.max_row + 1):
        cnpj_raw = ws.cell(row=row_idx, column=col_cnpj).value
        cnpj = normalizar_cnpj(cnpj_raw) if cnpj_raw else ""
        if not cnpj:
            continue

        fatura = valores_por_cnpj.get(cnpj)
        if not fatura:
            stats["cnpj_sem_fatura_na_iugu"] += 1
            continue

        total_cents = int(
            fatura.get("total_paid_cents")
            or fatura.get("total_cents")
            or 0
        )
        if total_cents <= 0:
            stats["fatura_sem_valor"] += 1
            continue

        valor_br = format_cents_to_br(total_cents)

        # valor_fatura
        valor_atual = ws.cell(row=row_idx, column=col_valor).value
        tem_valor_atual = valor_atual not in (None, "", 0, "0", "0,00")
        if tem_valor_atual and preservar_existentes:
            stats["valor_preservado"] += 1
        else:
            ws.cell(row=row_idx, column=col_valor, value=valor_br)
            stats["valor_atualizado"] += 1
            alteracoes_log.append((cnpj, "valor_fatura", valor_atual, valor_br))

        # descricao_boleto (opcional)
        if preencher_descricao and col_desc_boleto:
            desc_atual = ws.cell(row=row_idx, column=col_desc_boleto).value
            tem_desc_atual = desc_atual not in (None, "", " ")
            if tem_desc_atual and preservar_existentes:
                stats["descricao_preservada"] += 1
            else:
                nova_desc = _pegar_descricao_item(fatura)
                if nova_desc:
                    ws.cell(row=row_idx, column=col_desc_boleto, value=nova_desc)
                    stats["descricao_atualizada"] += 1
                    alteracoes_log.append(
                        (cnpj, "descricao_boleto", desc_atual, nova_desc)
                    )

    if not dry_run:
        wb.save(caminho)
        logger.info(f"Planilha salva: {caminho}")

    return {"stats": dict(stats), "alteracoes": alteracoes_log}


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument("--days", type=int, default=120, help="Janela em dias (default: 120)")
    parser.add_argument(
        "--preencher-descricao",
        action="store_true",
        help="Preenche também descricao_boleto com a descrição do item da última fatura",
    )
    parser.add_argument(
        "--preservar",
        action="store_true",
        help="Preserva valores já preenchidos na planilha (default: sobrescreve)",
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--arquivo",
        type=Path,
        default=None,
        help=f"Caminho da planilha (default: {settings.planilha_empresas})",
    )
    args = parser.parse_args()

    caminho = args.arquivo or settings.planilha_empresas
    logger.info(f"Planilha alvo: {caminho}")

    # 1. Carrega CNPJs da planilha
    repo = EmpresasRepository(caminho=caminho)
    repo.carregar()
    cnpjs_alvo = {cnpj for cnpj in repo._empresas.keys()}
    logger.info(f"CNPJs a enriquecer: {len(cnpjs_alvo)}")
    if not cnpjs_alvo:
        print("⚠️  Planilha vazia — nada a fazer.")
        return

    # 2. Busca última fatura de cada
    valores = buscar_ultima_fatura_por_cnpj(cnpjs_alvo, dias=args.days)

    # 3. Relatório do que vai ser feito
    print(f"\n{'=' * 70}")
    print(f"📋 ÚLTIMAS FATURAS ENCONTRADAS (janela de {args.days} dias)")
    print(f"{'=' * 70}")
    for cnpj in sorted(cnpjs_alvo):
        emp = repo._empresas.get(cnpj)
        nome = emp.razao_social if emp else "—"
        fatura = valores.get(cnpj)
        if fatura:
            total_cents = int(
                fatura.get("total_paid_cents") or fatura.get("total_cents") or 0
            )
            valor_br = format_cents_to_br(total_cents)
            data = fatura.get("created_at", "—")
            status = fatura.get("status", "?")
            print(f"  ✓ {nome:<45} R$ {valor_br:<12} [{status:<8}] {data}")
        else:
            print(f"  ✗ {nome:<45} (nenhuma fatura nos últimos {args.days} dias)")
    print()

    # 4. Aplica no arquivo
    resultado = atualizar_planilha(
        caminho=caminho,
        valores_por_cnpj=valores,
        preservar_existentes=args.preservar,
        preencher_descricao=args.preencher_descricao,
        dry_run=args.dry_run,
    )

    print(f"{'=' * 70}")
    print(f"📊 RESUMO {'(DRY-RUN — nada gravado)' if args.dry_run else ''}")
    print(f"{'=' * 70}")
    for chave, valor in resultado["stats"].items():
        print(f"  {chave:<30} {valor}")
    print()

    if args.dry_run:
        print("Para aplicar de verdade, rode sem --dry-run")
    else:
        print(f"✅ Planilha atualizada: {caminho}")
        print("Abra no Excel e revise antes de agendar os boletos.")


if __name__ == "__main__":
    main()
