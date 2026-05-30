"""
Importa clientes das faturas recentes da Iugu para a planilha de empresas autorizadas.

Filtra:
- Apenas CNPJ (14 dígitos). CPFs são ignorados.
- Últimos 30 dias (configurável via --days)
- Status: "paid" e "pending" por padrão (configurável)

Cuidado: esse script **substitui** o conteúdo atual de dados da planilha
(preserva apenas o cabeçalho e a aba de instruções).

Uso:
    python scripts/import_clients_from_iugu.py
    python scripts/import_clients_from_iugu.py --days 60
    python scripts/import_clients_from_iugu.py --status paid
    python scripts/import_clients_from_iugu.py --dry-run
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Iterable

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from loguru import logger
from openpyxl import load_workbook

from src.config import settings
from src.iugu_client import IuguClient
from src.spreadsheet import COLUNAS, Empresa, criar_planilha_modelo, normalizar_cnpj

CNPJ_LENGTH = 14  # empresa = 14 dígitos, CPF = 11


def _somente_digitos(valor) -> str:
    if not valor:
        return ""
    return "".join(filter(str.isdigit, str(valor)))


def _limpar(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _invoice_to_empresa(inv: dict) -> Empresa | None:
    """Converte uma fatura da Iugu em Empresa. Retorna None se não for CNPJ válido."""
    doc_raw = inv.get("payer_cpf_cnpj") or ""
    cnpj = normalizar_cnpj(doc_raw)
    if len(cnpj) != CNPJ_LENGTH:
        return None  # pula CPF ou CNPJ malformado

    return Empresa(
        cnpj=cnpj,
        razao_social=_limpar(inv.get("payer_name")),
        email=_limpar(inv.get("payer_email")),
        logradouro=_limpar(inv.get("payer_address_street")),
        numero=_limpar(inv.get("payer_address_number")),
        complemento=_limpar(inv.get("payer_address_complement")),
        bairro=_limpar(inv.get("payer_address_district")),
        cidade=_limpar(inv.get("payer_address_city")),
        uf=_limpar(inv.get("payer_address_state")).upper()[:2],
        cep=_somente_digitos(inv.get("payer_address_zip_code")),
        codigo_servico="",        # preencher manualmente por cliente
        descricao_servico="",     # preencher manualmente por cliente
        aliquota_iss=0.0,         # preencher manualmente por cliente
        observacoes=f"Importado da Iugu em {date.today().isoformat()}",
        ativo=True,
    )


def coletar_clientes_unicos(
    dias: int = 30,
    statuses: Iterable[str] = ("paid", "pending"),
) -> list[Empresa]:
    """Busca faturas, filtra CNPJs e deduplica."""
    data_fim = date.today()
    data_inicio = data_fim - timedelta(days=dias)
    logger.info(f"Período: {data_inicio} a {data_fim}")

    por_cnpj: dict[str, Empresa] = {}
    total_faturas_vistas = 0
    total_cpfs_pulados = 0

    with IuguClient() as client:
        for status in statuses:
            logger.info(f"→ Buscando faturas status='{status}'...")
            start = 0
            limit = 100
            pagina = 0
            while True:
                pagina += 1
                result = client.list_invoices(
                    status=status,
                    limit=limit,
                    start=start,
                    created_at_from=data_inicio.isoformat(),
                    created_at_to=data_fim.isoformat(),
                )
                items = result.get("items", []) or []
                total = result.get("totalItems", 0)
                logger.info(f"   Página {pagina}: {len(items)} faturas (de {total} no total)")

                for inv in items:
                    total_faturas_vistas += 1
                    emp = _invoice_to_empresa(inv)
                    if emp is None:
                        total_cpfs_pulados += 1
                        continue
                    # Mantém o primeiro encontrado (não sobrescreve se já existe)
                    if emp.cnpj not in por_cnpj:
                        por_cnpj[emp.cnpj] = emp

                if len(items) < limit:
                    break
                start += limit

    logger.info(
        f"✔ Total de faturas varridas: {total_faturas_vistas} "
        f"| CPFs pulados: {total_cpfs_pulados} "
        f"| Clientes únicos (CNPJ): {len(por_cnpj)}"
    )
    return list(por_cnpj.values())


def gravar_planilha(clientes: list[Empresa], caminho: Path) -> None:
    """Recria a planilha do zero com os clientes importados.

    Preserva o cabeçalho estilizado e a aba de instruções (via criar_planilha_modelo),
    mas limpa qualquer dado anterior (exemplos fictícios).
    """
    criar_planilha_modelo(caminho=caminho, com_exemplos=False)

    wb = load_workbook(caminho)
    ws = wb["Empresas Autorizadas"]

    clientes_ordenados = sorted(
        clientes, key=lambda c: (c.razao_social or "").lower()
    )

    for row_idx, emp in enumerate(clientes_ordenados, start=2):
        for col_idx, coluna in enumerate(COLUNAS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=getattr(emp, coluna))

    wb.save(caminho)
    logger.info(f"Planilha salva com {len(clientes)} clientes em {caminho}")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--days", type=int, default=30, help="Janela em dias (default: 30)")
    parser.add_argument(
        "--status",
        nargs="+",
        default=["paid", "pending"],
        help="Status de fatura a considerar (default: paid pending)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Apenas lista o que seria importado, sem modificar a planilha",
    )
    parser.add_argument(
        "--saida",
        type=Path,
        default=None,
        help="Caminho de sa\u00edda (default: settings.planilha_empresas)",
    )
    args = parser.parse_args()

    clientes = coletar_clientes_unicos(dias=args.days, statuses=args.status)

    if not clientes:
        logger.warning("Nenhum cliente encontrado — planilha n\u00e3o modificada.")
        return

    print(f"\n{'=' * 60}")
    print(f"📊 {len(clientes)} CLIENTES ÚNICOS ENCONTRADOS")
    print(f"{'=' * 60}\n")
    for emp in sorted(clientes, key=lambda c: (c.razao_social or "").lower()):
        cidade_uf = f"{emp.cidade}/{emp.uf}" if emp.cidade else "—"
        print(f"  • {emp.razao_social:<45} {emp.cnpj}  {cidade_uf}")
    print()

    if args.dry_run:
        logger.info("DRY RUN — planilha não foi modificada.")
        return

    caminho = Path(args.saida) if args.saida else settings.planilha_empresas
    gravar_planilha(clientes, caminho)

    print(f"\n✅ Planilha atualizada: {caminho}")
    print("Próximos passos:")
    print("  1. Abra a planilha e preencha codigo_servico, descricao_servico e aliquota_iss por cliente")
    print("  2. Marque ativo=False para clientes que NÃO devem receber NFS-e automática")
    print("  3. Revise os endereços (alguns podem vir incompletos da Iugu)")


if __name__ == "__main__":
    main()
