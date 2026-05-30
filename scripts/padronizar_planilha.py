"""
Padroniza os valores da planilha de empresas.

Corrige inconsistências de formatação:
  - Booleanos: converte True/False/VERDADEIRO/FALSO/bool → "Sim"/"Não"
  - valor_fatura: converte para formato "R$ 1.850,00"
  - Caixa: padroniza razao_social para UPPER CASE

IDEMPOTENTE: pode ser rodado várias vezes sem problema.

Uso:
    python scripts/padronizar_planilha.py
    python scripts/padronizar_planilha.py --dry-run
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from loguru import logger
from openpyxl import load_workbook

from src.config import settings
from src.spreadsheet import format_cents_to_br, parse_valor_br_to_cents

# Colunas booleanas da planilha
COLUNAS_BOOL = ["emitir_nf", "nf_na_criacao", "ativo"]


def _to_bool_str(valor) -> str | None:
    """Converte qualquer valor boolean-like para 'Sim' ou 'Não'. None se não é bool."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, bool):
        return "Sim" if valor else "Não"
    if isinstance(valor, str):
        v = valor.strip().lower()
        if v in ("true", "sim", "s", "1", "yes", "y", "verdadeiro"):
            return "Sim"
        if v in ("false", "não", "nao", "n", "0", "no", "falso"):
            return "Não"
    return None


def _formatar_valor_br(valor) -> str | None:
    """Converte valor para formato 'R$ 1.850,00'. Retorna None se vazio."""
    if valor is None or valor == "":
        return None
    cents = parse_valor_br_to_cents(valor)
    if cents <= 0:
        return None
    return f"R$ {format_cents_to_br(cents)}"


def fazer_backup(caminho: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = caminho.with_name(f"{caminho.stem}.bak_{ts}{caminho.suffix}")
    shutil.copy2(caminho, backup)
    logger.info(f"Backup criado: {backup.name}")
    return backup


def padronizar(caminho: Path, dry_run: bool = False) -> dict:
    """Padroniza a planilha. Retorna resumo das alterações."""
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    wb = load_workbook(caminho)

    # Procura aba de empresas
    aba = None
    for nome in wb.sheetnames:
        if "empresa" in nome.lower():
            aba = nome
            break
    if aba is None:
        aba = wb.sheetnames[0]
    ws = wb[aba]

    # Mapeia header
    header = {}
    for ci, cell in enumerate(ws[1], 1):
        if cell.value:
            header[str(cell.value).strip().lower()] = ci

    total_dados = ws.max_row - 1
    alteracoes = {"booleanos": 0, "valor_fatura": 0, "total_linhas": total_dados}
    detalhes = []

    for r in range(2, ws.max_row + 1):
        # Ignora linhas sem CNPJ
        col_cnpj = header.get("cnpj")
        if col_cnpj:
            cnpj_val = ws.cell(row=r, column=col_cnpj).value
            if not cnpj_val:
                continue

        razao_col = header.get("razao_social")
        razao = ws.cell(row=r, column=razao_col).value if razao_col else "?"

        # --- Padroniza booleanos ---
        for col_name in COLUNAS_BOOL:
            col_idx = header.get(col_name)
            if col_idx is None:
                continue

            cell = ws.cell(row=r, column=col_idx)
            valor_atual = cell.value
            novo = _to_bool_str(valor_atual)

            if novo is None:
                continue

            # Verifica se precisa atualizar
            if isinstance(valor_atual, str) and valor_atual.strip() == novo:
                continue  # Já está no padrão

            detalhes.append(
                f"  Linha {r} ({razao}): {col_name} = {valor_atual!r} → {novo!r}"
            )
            if not dry_run:
                cell.value = novo
            alteracoes["booleanos"] += 1

        # --- Padroniza valor_fatura ---
        col_valor = header.get("valor_fatura")
        if col_valor:
            cell_valor = ws.cell(row=r, column=col_valor)
            valor_atual = cell_valor.value
            if valor_atual is not None and str(valor_atual).strip():
                novo_valor = _formatar_valor_br(valor_atual)
                if novo_valor and str(valor_atual).strip() != novo_valor:
                    detalhes.append(
                        f"  Linha {r} ({razao}): valor_fatura = {valor_atual!r} → {novo_valor!r}"
                    )
                    if not dry_run:
                        cell_valor.value = novo_valor
                    alteracoes["valor_fatura"] += 1

    if not dry_run and (alteracoes["booleanos"] > 0 or alteracoes["valor_fatura"] > 0):
        fazer_backup(caminho)
        wb.save(caminho)
        logger.info(f"✅ Planilha padronizada: {caminho}")
    else:
        if dry_run:
            logger.info("DRY RUN — nenhuma alteração feita.")
        else:
            logger.info("✅ Planilha já está padronizada. Nada a fazer.")

    alteracoes["detalhes"] = detalhes
    return alteracoes


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument(
        "--arquivo", type=Path, default=None,
        help=f"Caminho da planilha (default: {settings.planilha_empresas})",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    caminho = args.arquivo or settings.planilha_empresas
    logger.info(f"Planilha alvo: {caminho}")

    try:
        resultado = padronizar(caminho, dry_run=args.dry_run)
    except Exception as exc:
        logger.exception("Falha na padronização")
        print(f"\n❌ ERRO: {exc}")
        sys.exit(1)

    print(f"\n{'=' * 60}")
    modo = "DRY RUN" if args.dry_run else "PADRONIZAÇÃO"
    print(f"📊 {modo} — Resumo")
    print(f"{'=' * 60}")
    print(f"  Linhas de dados: {resultado['total_linhas']}")
    print(f"  Booleanos corrigidos: {resultado['booleanos']}")
    print(f"  Valores formatados: {resultado['valor_fatura']}")

    if resultado.get("detalhes"):
        print(f"\nAlterações:")
        for d in resultado["detalhes"]:
            print(d)
    else:
        print("\n  ✅ Tudo já está no padrão!")


if __name__ == "__main__":
    main()
