"""
Simplifica o schema da planilha:
- REMOVE colunas de endereço (logradouro, numero, complemento, bairro,
  cidade, uf, cep) — endereço agora vem da Iugu automaticamente.
- ADICIONA coluna `emitir_nf` (bool) — controla se a NFS-e deve ser emitida
  automaticamente ao receber pagamento.

IDEMPOTENTE: pode ser rodado várias vezes sem problema.
PRESERVA todos os dados existentes (CNPJs, serviço, boleto recorrente).

Uso:
    python scripts/migrate_simplify_schema.py
    python scripts/migrate_simplify_schema.py --dry-run
    python scripts/migrate_simplify_schema.py --emitir-nf-default False
    python scripts/migrate_simplify_schema.py --arquivo outra_planilha.xlsx

⚠️ Feche o Excel antes de rodar.
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import settings

# Colunas de endereço que serão REMOVIDAS
COLUNAS_ENDERECO = [
    "logradouro", "numero", "complemento", "bairro",
    "cidade", "uf", "cep",
]

# Nova coluna que será ADICIONADA (se não existir)
COLUNA_NOVA = "emitir_nf"
# Posição alvo: logo depois de "aliquota_iss"
COLUNA_ANCORA_ANTES = "aliquota_iss"

LARGURAS = {"emitir_nf": 12}


def fazer_backup(caminho: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = caminho.with_name(f"{caminho.stem}.bak_{ts}{caminho.suffix}")
    shutil.copy2(caminho, backup)
    logger.info(f"Backup criado: {backup.name}")
    return backup


def migrar(
    caminho: Path,
    emitir_nf_default: bool = True,
    dry_run: bool = False,
) -> dict:
    """Aplica a migração. Retorna dict com o que foi feito."""
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    wb = load_workbook(caminho)
    if "Empresas Autorizadas" not in wb.sheetnames:
        raise ValueError(
            f'Aba "Empresas Autorizadas" não encontrada. Abas: {wb.sheetnames}'
        )

    ws = wb["Empresas Autorizadas"]

    # Mapeia colunas existentes
    header_atual = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            header_atual[str(cell.value).strip().lower()] = col_idx

    colunas_endereco_presentes = [c for c in COLUNAS_ENDERECO if c in header_atual]
    coluna_nova_presente = COLUNA_NOVA in header_atual
    total_dados = ws.max_row - 1

    if not colunas_endereco_presentes and coluna_nova_presente:
        logger.info("✅ Planilha já está no schema atualizado. Nada a fazer.")
        return {"status": "ja_migrada", "linhas_dados": total_dados}

    logger.info(
        f"Colunas de endereço a remover: {colunas_endereco_presentes or '—'}"
    )
    logger.info(
        f"Coluna '{COLUNA_NOVA}': {'já existe' if coluna_nova_presente else 'será adicionada'}"
    )

    if dry_run:
        logger.info("DRY RUN — nenhuma alteração será feita.")
        return {
            "status": "dry_run",
            "colunas_endereco_a_remover": colunas_endereco_presentes,
            "adicionaria_emitir_nf": not coluna_nova_presente,
            "linhas_dados": total_dados,
        }

    # Backup antes de mexer
    fazer_backup(caminho)

    # --- PASSO 1: remover colunas de endereço ---
    # openpyxl delete_cols opera por índice; se removermos em ordem decrescente
    # os índices das colunas à esquerda não se alteram.
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    # Re-mapeia o header a cada iteração (índices mudam após delete)
    def _mapear_header():
        m = {}
        for ci, c in enumerate(ws[1], start=1):
            if c.value:
                m[str(c.value).strip().lower()] = ci
        return m

    removidas = []
    for nome in COLUNAS_ENDERECO:
        header = _mapear_header()
        idx = header.get(nome)
        if idx is None:
            continue
        ws.delete_cols(idx, amount=1)
        removidas.append(nome)
        logger.debug(f"Coluna removida: {nome} (era posição {idx})")

    # --- PASSO 2: adicionar emitir_nf (se ainda não existir) ---
    header = _mapear_header()
    adicionada = False
    if COLUNA_NOVA not in header:
        ancora = header.get(COLUNA_ANCORA_ANTES)
        if ancora is None:
            raise ValueError(
                f"Coluna âncora '{COLUNA_ANCORA_ANTES}' não encontrada após remoção. "
                f"Colunas atuais: {list(header.keys())}"
            )
        pos = ancora + 1
        ws.insert_cols(pos, amount=1)
        cell = ws.cell(row=1, column=pos, value=COLUNA_NOVA)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(pos)].width = LARGURAS[COLUNA_NOVA]

        # Preenche default para todas as linhas de dados
        default_str = "Sim" if emitir_nf_default else "Não"
        for r in range(2, ws.max_row + 1):
            # Só preenche linhas que têm CNPJ (não mexe em linhas vazias/cabeçalho)
            header_cnpj = _mapear_header().get("cnpj")
            if header_cnpj and ws.cell(row=r, column=header_cnpj).value:
                ws.cell(row=r, column=pos, value=default_str)
        adicionada = True
        logger.info(
            f"Coluna '{COLUNA_NOVA}' adicionada na posição "
            f"{get_column_letter(pos)} com default {default_str} para {total_dados} linhas"
        )

    # Atualiza aba de instruções
    _atualizar_aba_instrucoes(wb)

    wb.save(caminho)
    logger.info(f"✅ Planilha migrada: {caminho}")

    return {
        "status": "migrada",
        "colunas_endereco_removidas": removidas,
        "coluna_emitir_nf_adicionada": adicionada,
        "emitir_nf_default": emitir_nf_default,
        "linhas_dados": total_dados,
    }


def _atualizar_aba_instrucoes(wb) -> None:
    """Recria a aba 'Instruções' com o texto atualizado."""
    if "Instruções" in wb.sheetnames:
        del wb["Instruções"]
    ws_info = wb.create_sheet("Instruções")
    instrucoes = [
        ("COMO USAR ESTA PLANILHA", True),
        ("", False),
        ("1. Preencha uma linha por empresa cadastrada.", False),
        ("2. O CNPJ é a chave de busca — com ou sem formatação.", False),
        ("3. Endereço NÃO é mais necessário aqui — vem automaticamente da Iugu.", False),
        ("", False),
        ("CAMPOS DE NFS-e:", True),
        ("4. codigo_servico: https://iss.fazenda.df.gov.br/online", False),
        ("5. aliquota_iss: percentual (ex: 2.0 para 2%)", False),
        ("6. emitir_nf: True = emite NFS-e ao receber pagamento | False = não emite.", False),
        ("", False),
        ("CAMPOS DE BOLETO RECORRENTE:", True),
        ("7. descricao_boleto: texto que aparece no boleto.", False),
        ("8. valor_fatura: formato BR (1850,00). Vazio = sem cobrança automática.", False),
        ("9. dia_criacao_fatura: dia do mês (1-31). Vencimento: 10 dias após.", False),
        ("", False),
        ("CONTROLE GERAL:", True),
        ("10. ativo: False desabilita TUDO.", False),
        ("    Para desligar só a NFS-e mantendo boleto: use emitir_nf=False.", False),
        ("    Para desligar só o boleto mantendo NFS-e: deixe valor/dia em branco.", False),
    ]
    for idx, (texto, negrito) in enumerate(instrucoes, start=1):
        cell = ws_info.cell(row=idx, column=1, value=texto)
        if negrito:
            cell.font = Font(bold=True, size=12)
    ws_info.column_dimensions["A"].width = 90


def _parse_bool(valor: str) -> bool:
    return valor.strip().lower() in ("true", "sim", "s", "1", "yes", "y", "t")


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument(
        "--arquivo",
        type=Path,
        default=None,
        help=f"Caminho da planilha (default: {settings.planilha_empresas})",
    )
    parser.add_argument(
        "--emitir-nf-default",
        type=_parse_bool,
        default=True,
        help="Valor default para emitir_nf em linhas existentes (default: True)",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    caminho = args.arquivo or settings.planilha_empresas
    logger.info(f"Planilha alvo: {caminho}")

    try:
        resultado = migrar(
            caminho,
            emitir_nf_default=args.emitir_nf_default,
            dry_run=args.dry_run,
        )
    except Exception as exc:
        logger.exception("Falha na migração")
        print(f"\n❌ ERRO: {exc}")
        sys.exit(1)

    print(f"\n{'=' * 60}")
    print(f"📊 MIGRAÇÃO — status: {resultado['status'].upper()}")
    print(f"{'=' * 60}")
    if resultado["status"] == "migrada":
        print(f"Colunas de endereço removidas: {resultado['colunas_endereco_removidas']}")
        print(f"Coluna 'emitir_nf' adicionada: {resultado['coluna_emitir_nf_adicionada']}")
        print(f"Default emitir_nf: {resultado['emitir_nf_default']}")
        print(f"Linhas preservadas: {resultado['linhas_dados']}")
    elif resultado["status"] == "dry_run":
        print(f"Seriam removidas: {resultado['colunas_endereco_a_remover']}")
        print(f"Adicionaria emitir_nf: {resultado['adicionaria_emitir_nf']}")
        print(f"Linhas de dados: {resultado['linhas_dados']}")
    elif resultado["status"] == "ja_migrada":
        print(f"Nada a fazer. Linhas de dados: {resultado['linhas_dados']}")


if __name__ == "__main__":
    main()
